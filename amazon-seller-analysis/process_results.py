#!/usr/bin/env python3
"""
Amazon セラーリサーチ 利益計算 & Excel出力スクリプト

使い方:
  python3 process_results.py --json '<SS_EXTRACT_RESULTのJSON>'
  または
  python3 process_results.py --file data.json
  または（インタラクティブ）
  python3 process_results.py

出力: セラーリサーチリスト.xlsx（実利益率降順）
"""

import json
import sys
import os
import re
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 利益計算ロジック
# ============================================================
COST_RATE = 0.20       # 原価率 = 価格の20%
MIN_NET_MARGIN = 0.30  # 最終フィルタ: 実利益率30%以上

def calculate_profit(price: float, selling_fee: float, fba_shipping: float) -> dict:
    """
    実利益計算
    実利益 = 価格 - 販売手数料 - FBA配送料 - 原価(価格×20%)
    実利益率 = 実利益 ÷ 価格
    """
    cost = price * COST_RATE
    net_profit = price - selling_fee - fba_shipping - cost
    net_margin = net_profit / price if price > 0 else 0
    return {
        'cost': round(cost),
        'net_profit': round(net_profit),
        'net_margin': round(net_margin, 4),
        'net_margin_pct': round(net_margin * 100, 1),
    }


def estimate_fees_from_fba_total(price: float, fba_total: float) -> tuple[float, float]:
    """
    FBA合計手数料から販売手数料とFBA配送料を推定分割
    Amazon JP: 販売手数料 = 価格 × 8〜10%（カテゴリ依存）
    残りをFBA配送料とする（簡易推定）
    """
    # 販売手数料: 参照価格ベースで8%と仮定（ペット用品カテゴリ）
    selling_fee = round(price * 0.08)
    fba_shipping = max(0, fba_total - selling_fee)
    return selling_fee, fba_shipping


# ============================================================
# データ処理
# ============================================================
def process_products(raw_products: list) -> list:
    """一次フィルタ済み商品を詳細計算して最終候補を返す"""
    results = []

    for p in raw_products:
        price = float(p.get('price', 0))
        fba_total = float(p.get('fba_fee', 0))
        monthly_sales = float(p.get('monthly_sales', 0))
        sales_count = float(p.get('sales_count', 0))

        if price < 950:
            continue

        # FBA手数料の詳細が別途取得済みの場合はそちらを使用
        selling_fee = float(p.get('selling_fee', 0))
        fba_shipping = float(p.get('fba_shipping', 0))

        if selling_fee == 0 and fba_shipping == 0 and fba_total > 0:
            selling_fee, fba_shipping = estimate_fees_from_fba_total(price, fba_total)
        elif selling_fee == 0:
            # FBA合計もない場合はデフォルト推定
            selling_fee = round(price * 0.08)
            fba_shipping = round(price * 0.12)  # 一般的なFBA配送料推定

        calc = calculate_profit(price, selling_fee, fba_shipping)

        # 最終フィルタ: 実利益率30%以上
        if calc['net_margin'] < MIN_NET_MARGIN:
            continue

        results.append({
            '商品名': p.get('name', ''),
            '価格': price,
            '販売手数料': selling_fee,
            'FBA配送料': fba_shipping,
            '原価(20%)': calc['cost'],
            '実利益': calc['net_profit'],
            '実利益率': calc['net_margin_pct'],
            '月商': monthly_sales,
            '販売数(30日)': sales_count,
            'レビュー数': p.get('reviews', 0),
            'セラー数': p.get('sellers', 0),
            'バリエーション数': p.get('variations', 0),
            '粗利益率(SS)': p.get('gross_margin', 0),
            'ASIN': p.get('asin', ''),
            'URL': p.get('url', ''),
            'ブランド': p.get('brand', ''),
            'カテゴリ': p.get('category', ''),
            '備考': '手数料推定値' if p.get('selling_fee', 0) == 0 else '',
        })

    # 実利益率降順でソート
    results.sort(key=lambda x: x['実利益率'], reverse=True)
    return results


# ============================================================
# Excel出力
# ============================================================
COLUMN_ORDER = [
    '商品名', '価格', '販売手数料', 'FBA配送料', '原価(20%)',
    '実利益', '実利益率', '月商', '販売数(30日)', 'レビュー数',
    'セラー数', 'バリエーション数', 'URL',
]

COLUMN_WIDTHS = {
    '商品名': 50, '価格': 10, '販売手数料': 12, 'FBA配送料': 12,
    '原価(20%)': 10, '実利益': 10, '実利益率': 10, '月商': 12,
    '販売数(30日)': 12, 'レビュー数': 10, 'セラー数': 10,
    'バリエーション数': 14, 'URL': 50,
}

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
GOOD_FILL  = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
GREAT_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

def write_excel(results: list, seller_name: str, output_path: str):
    """セラーリサーチリスト.xlsx に追記（シート名 = ショップ名）"""
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
        # デフォルトシートを削除
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    # 既存シートがあれば上書き
    if seller_name in wb.sheetnames:
        del wb[seller_name]

    ws = wb.create_sheet(title=seller_name[:31])  # シート名31文字制限

    # ヘッダー行
    for col_idx, col_name in enumerate(COLUMN_ORDER, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 15)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    # データ行
    for row_idx, item in enumerate(results, 2):
        for col_idx, col_name in enumerate(COLUMN_ORDER, 1):
            val = item.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center')

            # 数値フォーマット
            if col_name == '価格':
                cell.number_format = '¥#,##0'
            elif col_name in ('販売手数料', 'FBA配送料', '原価(20%)', '実利益', '月商'):
                cell.number_format = '¥#,##0'
            elif col_name == '実利益率':
                cell.value = val / 100  # パーセント値をExcel用に変換
                cell.number_format = '0.0%'
            elif col_name in ('販売数(30日)', 'レビュー数', 'セラー数', 'バリエーション数'):
                cell.number_format = '#,##0'
            elif col_name == 'URL':
                cell.alignment = Alignment(vertical='center', wrap_text=False)

        # 実利益率による行の色分け
        margin = item.get('実利益率', 0)
        row_fill = GREAT_FILL if margin >= 50 else (GOOD_FILL if margin >= 30 else None)
        if row_fill:
            for col_idx in range(1, len(COLUMN_ORDER) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # サマリー行
    summary_row = len(results) + 3
    ws.cell(row=summary_row, column=1, value=f'抽出日時: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    ws.cell(row=summary_row + 1, column=1, value=f'最終候補数: {len(results)}件（実利益率30%以上）')
    ws.cell(row=summary_row + 2, column=1, value='※販売手数料・FBA配送料は「推定値」を含む。電卓マークで取得した実値があれば上書きしてください。')

    wb.save(output_path)
    print(f"保存完了: {output_path} (シート: {seller_name})")
    return output_path


# ============================================================
# メイン
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='Amazon セラーリサーチ結果をExcel出力')
    parser.add_argument('--json', type=str, help='SS_EXTRACT_RESULT JSONデータ（文字列）')
    parser.add_argument('--file', type=str, help='JSONファイルパス')
    parser.add_argument('--seller', type=str, default='Petifam（ペティファーム）', help='ショップ名（シート名）')
    parser.add_argument('--output', type=str, default='/home/user/bellrock-rpp/amazon-seller-analysis/セラーリサーチリスト.xlsx',
                        help='出力Excelパス')
    args = parser.parse_args()

    # データ読み込み
    raw_data = None

    if args.json:
        # コマンドライン引数から（SS_EXTRACT_RESULT: プレフィックスを除去）
        json_str = args.json
        if json_str.startswith('SS_EXTRACT_RESULT:'):
            json_str = json_str[len('SS_EXTRACT_RESULT:'):]
        raw_data = json.loads(json_str)

    elif args.file:
        with open(args.file, 'r', encoding='utf-8') as f:
            content = f.read()
            # SS_EXTRACT_RESULT: プレフィックスを除去
            match = re.search(r'SS_EXTRACT_RESULT:(\[.*\])', content, re.DOTALL)
            if match:
                raw_data = json.loads(match.group(1))
            else:
                raw_data = json.loads(content)

    else:
        # インタラクティブモード
        print("コンソールからコピーした SS_EXTRACT_RESULT: 行を貼り付けてください。")
        print("（貼り付け後、空行でEnterを押してください）")
        lines = []
        while True:
            try:
                line = input()
                if not line and lines:
                    break
                lines.append(line)
            except EOFError:
                break
        content = '\n'.join(lines)
        match = re.search(r'SS_EXTRACT_RESULT:(\[.*\])', content, re.DOTALL)
        if match:
            raw_data = json.loads(match.group(1))
        else:
            try:
                raw_data = json.loads(content)
            except json.JSONDecodeError:
                print("エラー: 有効なJSONが見つかりません")
                sys.exit(1)

    if not raw_data:
        print("データが空です")
        sys.exit(1)

    print(f"\n入力データ: {len(raw_data)}件")

    # 処理
    results = process_products(raw_data)
    print(f"最終候補（実利益率30%以上）: {len(results)}件\n")

    if not results:
        print("該当商品なし。条件を満たす商品が見つかりませんでした。")
        print("確認事項:")
        print("  - 価格950円以上")
        print("  - 月商70,000〜500,000円")
        print("  - 実利益率30%以上（価格 - 販売手数料 - FBA配送料 - 原価20% ÷ 価格）")
        sys.exit(0)

    # 結果表示
    print(f"{'商品名':<40} {'価格':>8} {'実利益率':>8} {'月商':>10} {'レビュー':>6}")
    print('-' * 78)
    for r in results:
        name = r['商品名'][:38]
        print(f"{name:<40} {r['価格']:>7,.0f}円 {r['実利益率']:>6.1f}% {r['月商']:>9,.0f}円 {r['レビュー数']:>5}")

    # Excel出力
    output_path = write_excel(results, args.seller, args.output)
    print(f"\n✓ Excel出力完了: {output_path}")


if __name__ == '__main__':
    main()
