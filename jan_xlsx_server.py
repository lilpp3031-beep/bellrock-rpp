#!/usr/bin/env python3
"""
JAN登録ツール用 Excel生成サーバー (port 8082)
HTMLツールからPOSTされたJSONを受け取り、openpyxlでXLSXを生成して返す。
GS1テンプレートのメタデータ（drawings, sharedStrings, printerSettings）を完全保持。
"""
import json, io, os, sys
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse
import openpyxl

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__),
    '.playwright-mcp', 'gjdb-product-upload-20260502.xlsx')

# Downloadsに新テンプレートがあればそちらを優先
NEW_TEMPLATE = os.path.expanduser('~/Downloads/gjdb_product_uploadform.xlsx')
if os.path.exists(NEW_TEMPLATE):
    TEMPLATE_PATH = NEW_TEMPLATE

HEADERS = [
    'GTINステータスコード','GTINステータス','GTIN','GS1事業者コード','事業者名',
    '商品名','商品名（カナ）','取扱品目コード','取扱品目名称','JICFS分類コード',
    'JICFS分類名称','GPC（GS1商品分類）コード','GPC（GS1商品分類）名称','ブランド名',
    '内容量','内容量単位名称','内容量単位コード','表示用規格','商品名（詳細）',
    '消費者向け区分','自社商品コード','品名','商品情報URL','商品コメント',
    'サイズ（幅）','サイズ（幅）単位名称','サイズ（幅）単位コード',
    'サイズ（高さ）','サイズ（高さ）単位名称','サイズ（高さ）単位コード',
    'サイズ（奥行き）','サイズ（奥行き）単位名称','サイズ（奥行き）単位コード',
    '総重量','総重量単位名称','総重量単位コード','希望小売価格','オープン価格フラグ',
    '軽減標準判定区分','消費税区分','消費税率','原産国（地域）コード','原産国（地域）',
    '販売対象国（地域）コード','販売対象国（地域）',
    '情報公開日','出荷可能日','出荷終了日','GTIN使用終了日',
    '言語コード1','言語1','他言語商品情報URL1',
    '言語コード2','言語2','他言語商品情報URL2',
    '言語コード3','言語3','他言語商品情報URL3',
    '言語コード4','言語4','他言語商品情報URL4',
    '言語コード5','言語5','他言語商品情報URL5',
    '登録事業者用メモ','登録日時','更新日時',
]

def col(name):
    return HEADERS.index(name) + 1  # 1-indexed for openpyxl

def generate_xlsx(payload):
    groups = payload.get('groups', [])
    selected_gs1 = payload.get('selectedGs1', '')

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['商品情報リスト']

    def to_full_width(text):
        """半角英数記号→全角、半角スペース→全角スペース"""
        if not text: return ''
        result = []
        for c in text:
            code = ord(c)
            if 0x21 <= code <= 0x7E:        # 半角英数記号
                result.append(chr(code + 0xFEE0))
            elif c == ' ':                   # 半角スペース
                result.append('　')
            else:
                result.append(c)
        return ''.join(result)

    def fix_kana(text):
        """カナ文字列の正規化：制御文字除去・半角スペース→全角・100文字以内"""
        if not text: return ''
        import re
        text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)  # 制御文字除去
        text = text.replace(' ', '　')                   # 半角スペース→全角
        return text[:100]                                     # 100文字以内

    row = 4  # データは4行目から
    for g in groups:
        brand = g.get('fields', {}).get('brandCustom') or g.get('fields', {}).get('brand', '')
        product_name = to_full_width(g.get('name', '').strip())
        fields = g.get('fields', {})

        for v in g.get('variants', []):
            color = to_full_width(v.get('color', ''))
            size  = to_full_width(v.get('size', ''))
            spec  = '　'.join(filter(None, [color, size]))
            final_name = product_name + ('　' + spec if spec else '')
            raw_detail = v.get('detail', '') if v.get('detailEdited') else \
                '　'.join(filter(None, [brand, product_name, color, size]))
            detail = to_full_width(raw_detail)
            jan  = v.get('jan', '')
            sku  = v.get('sku', '')

            def s(name, val):
                if val is None or val == '':
                    return
                ws.cell(row=row, column=col(name)).value = str(val)

            def n(name, val):
                try:
                    ws.cell(row=row, column=col(name)).value = int(val)
                except (ValueError, TypeError):
                    pass

            s('GTINステータスコード', '02')
            s('GTIN', jan)
            s('GS1事業者コード', selected_gs1)
            s('商品名', final_name)
            s('商品名（カナ）', fix_kana(fields.get('nameKana', '')))
            if fields.get('itemCode'): n('取扱品目コード', fields['itemCode'])
            if fields.get('gpc'):      n('GPC（GS1商品分類）コード', fields['gpc'])
            if fields.get('jicfs'):    n('JICFS分類コード', fields['jicfs'])
            s('ブランド名', brand)
            if fields.get('qty'):
                try:
                    ws.cell(row=row, column=col('内容量')).value = float(fields['qty'])
                except ValueError:
                    pass
            s('内容量単位名称', fields.get('qtyUnit', '個'))
            s('内容量単位コード', '001')
            s('表示用規格', spec)
            s('商品名（詳細）', detail)
            s('消費者向け区分', fields.get('consumer', '1'))
            if sku: s('自社商品コード', sku)
            s('オープン価格フラグ', '1')
            s('原産国（地域）コード', fields.get('origin', '156'))
            s('販売対象国（地域）コード', '392')
            # 言語コード1は設定しない（URLなしで設定するとGS1エラーになるため）

            row += 1

    # openpyxlで一旦保存
    tmp = io.BytesIO()
    wb.save(tmp)
    tmp.seek(0)

    # 元テンプレートから欠落ファイルをコピーして完全なxlsxに再構築
    import zipfile
    result = io.BytesIO()
    with zipfile.ZipFile(tmp, 'r') as gen_zip, \
         zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl_zip, \
         zipfile.ZipFile(result, 'w', zipfile.ZIP_DEFLATED) as out_zip:

        gen_names = set(gen_zip.namelist())
        tmpl_names = set(tmpl_zip.namelist())

        # 生成ファイルの全エントリを書き込む
        for name in gen_zip.namelist():
            out_zip.writestr(name, gen_zip.read(name))

        # 元テンプレートにあって生成ファイルに無いものをコピー
        for name in tmpl_names - gen_names:
            out_zip.writestr(name, tmpl_zip.read(name))

    return result.getvalue(), row - 4  # bytesと件数


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass  # ログ抑制

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        if urlparse(self.path).path != '/generate':
            self.send_response(404); self.end_headers(); return

        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length)
        try:
            payload = json.loads(body)
            xlsx_bytes, count = generate_xlsx(payload)
            from datetime import date
            fname = f'gjdb_product_upload_{date.today().strftime("%Y%m%d")}.xlsx'
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)
            print(f'[OK] {count}件のExcelを生成', flush=True)
        except Exception as e:
            import traceback
            msg = traceback.format_exc().encode()
            self.send_response(500)
            self._cors()
            self.send_header('Content-Type', 'text/plain')
            self.send_header('Content-Length', str(len(msg)))
            self.end_headers()
            self.wfile.write(msg)
            print(f'[ERROR] {e}', flush=True)

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')


if __name__ == '__main__':
    port = 8082
    print(f'JAN Excel生成サーバー起動: http://localhost:{port}', flush=True)
    print(f'テンプレート: {TEMPLATE_PATH}', flush=True)
    HTTPServer(('localhost', port), Handler).serve_forever()
